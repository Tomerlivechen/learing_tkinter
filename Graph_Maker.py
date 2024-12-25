# -*- coding: utf-8 -*-
"""
Created on Wed Mar 27 12:10:02 2024

@author: tomer
"""
# where pip
# python -m ensurepip
# pip install pandas numpy scipy plotly openpyxl requests datetime kaleido scikit_posthocs
import requests
import pandas as pd
import numpy as np
from scipy.stats import sem
import plotly.graph_objects as go
from scipy.stats import tukey_hsd
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import os
from openpyxl import load_workbook
from datetime import date
from scikit_posthocs import posthoc_tukey

current_date = date.today()
current_year = current_date.year


class DataSet:
    def __init__(self, x_data, y_data, color, export_folder, x_axis="", y_axis="", title="", p_value=[]):
        self.x_data = x_data
        self.y_data = y_data
        self.color = color
        self.x_axis = x_axis
        self.y_axis = y_axis
        self.title = title
        self.p_value = p_value
        self.export_folder = export_folder
        self.width_value, self.height_value, self.title_font_size, self.legend_font_size = getQuality()


def barplot(dataSet):
    y_data_average = averageArrayFromMatrix(
        transpose_matrix_right(dataSet.y_data))
    error_y = standartErrorArrayFromMatrix(dataSet.y_data)

    # Create traces for each dataset
    traces = []
    for i in range(len(dataSet.x_data)):
        trace = go.Bar(
            x=[dataSet.x_data[i]],
            y=[y_data_average[i]],
            marker_color=dataSet.color[i],
            error_y=dict(type='data', array=[error_y[i]]),
            name=dataSet.x_data[i])
        traces.append(trace)

    layout = go.Layout(
        title=dict(text=dataSet.title, x=0.5, font=dict(
            size=dataSet.title_font_size)),
        xaxis=dict(title=dataSet.x_axis, title_font=dict(
            size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
        yaxis=dict(title=dataSet.y_axis, title_font=dict(
            size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
        legend=dict(font=dict(size=dataSet.legend_font_size),
                    yanchor="top", y=1, xanchor="left", x=1),
        barmode='group',
        width=dataSet.width_value,
        height=dataSet.height_value)
    fig = go.Figure(data=traces, layout=layout)

    return fig


def stripplot(dataSet):
    y_data_average = averageArrayFromMatrix(
        transpose_matrix_right(dataSet.y_data))
    traces = []
    for i, dataset in enumerate(dataSet.x_data):

        scatter_data = go.Scatter(x=[dataset] * len(transpose_matrix_right(dataSet.y_data)[i]),
                                  y=transpose_matrix_right(dataSet.y_data)[i],
                                  mode='markers',
                                  marker=dict(
                                      color=dataSet.color[i], size=dataSet.legend_font_size-5),
                                  name=f"{dataset}"
                                  )
        traces.append(scatter_data)

    scatter_data = go.Scatter(x=dataSet.x_data,
                              y=y_data_average,
                              mode='markers',
                              marker=dict(symbol='line-ew-open',
                                          size=dataSet.title_font_size, color="Black"),
                              name='Mean')
    traces.append(scatter_data)

    # Update layout
    layout = go.Layout(title=dict(text=dataSet.title, x=0.5, font=dict(size=dataSet.title_font_size)),
                       xaxis=dict(title=dataSet.x_axis, title_font=dict(
                           size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
                       yaxis=dict(title=dataSet.y_axis, title_font=dict(
                           size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
                       legend=dict(font=dict(size=dataSet.legend_font_size)),
                       width=dataSet.width_value,
                       height=dataSet.height_value)

    fig = go.Figure(data=traces, layout=layout)

    return fig


def boxplot(dataSet):
    fig = go.Figure()
    for i, dataset in enumerate(dataSet.x_data):
        boxdata = go.Box(
            y=transpose_matrix_right(dataSet.y_data)[i], fillcolor=dataSet.color[i], name=f"{dataSet.x_data[i]}", line_color='black')
        fig.add_trace(boxdata)
    fig.update_layout(title=dict(text=dataSet.title, x=0.5, font=dict(size=dataSet.title_font_size)),
                      xaxis=dict(title=dataSet.x_axis, title_font=dict(
                          size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
                      yaxis=dict(title=dataSet.y_axis, title_font=dict(
                          size=dataSet.title_font_size), tickfont=dict(size=dataSet.legend_font_size)),
                      legend=dict(font=dict(size=dataSet.legend_font_size)),
                      width=dataSet.width_value,
                      height=dataSet.height_value)
    return fig


def licenseing():
    url = "https://api.github.com/users/tomerlivechen"
    response = requests.get(url)
    if response.status_code == 200:
        user_data = response.json()
        return int(user_data["created_at"][0:4])
    else:
        return 2022


def averageArrayFromMatrix(matrix):
    array = np.zeros(len(matrix))
    for i in range(len(matrix)):
        array[i] = averageArray(matrix[i])
    return array


def standartErrorArrayFromMatrix(matrix):
    n = len(matrix)
    array = [None] * n
    for i in range(n):
        valid_values = [val for val in matrix[i] if val is not None]
        if len(valid_values) > 1:
            array[i] = sem(valid_values)
        else:
            array[i] = None
    return array


def averageArray(array):
    filtered_array = [val for val in array if val is not None]
    if len(filtered_array) == 0:
        return None
    return sum(filtered_array) / len(filtered_array)


def transpose_matrix_right(matrix):
    transposed_matrix = list(zip(*matrix))
    rotated_matrix = [list(row[::-1]) for row in transposed_matrix]
    return rotated_matrix


def transpose_matrix_left(matrix):
    transposed_matrix = list(zip(*matrix))
    rotated_matrix = transposed_matrix[::-1]
    return rotated_matrix


def exportfigure(figure, filepath):
    figure.write_image(filepath)


def tukey_kermer(dataSet):
    try:
        max_length = max(len(row) for row in dataSet.y_data)
        padded_matrix = np.array(
            [row + [np.nan] * (max_length - len(row)) for row in dataSet.y_data])
        matrix = padded_matrix.T
        p_values = posthoc_tukey(matrix)
        return p_values
    except Exception as e:
        message = f"Error reading Excel file: data incompateble Sheet name incorrect \nReport:{e}"
        show_error_message(message)
        return None


def printmatrix(pval, dataSet):
    p_matrix = pval.values
    df = pd.DataFrame(p_matrix, index=dataSet.x_data,
                      columns=dataSet.x_data)
    df.to_csv(
        f"{dataSet.export_folder}/{dataSet.title} tukey kermer p-Value table.csv")


def generate_matrix(num):
    matrix = []
    for i in range(num):
        for j in range(num):
            if i != j:
                matrix.append([i, j])
    return matrix


def read_excel_data(file_path, sheet_name="Sheet1"):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        color = np.array(data[0])
        headers = np.array(data[1])

        matrix = []
        for row in data[2:]:
            numeric_row = [cell if cell is not None else None for cell in row]
            matrix.append(numeric_row)

        return headers, matrix, color

    except Exception as e:
        message = f"Error reading Excel file: {e}"
        show_error_message(message)
        return None, None


def DataSetConstructior(file_path, sheet_name, x_axis, y_axis, title, export_folder):
    [x_data, y_data, color] = read_excel_data(file_path, sheet_name)
    data_set = DataSet(x_data, y_data, color, export_folder, x_axis,
                       y_axis, title)
    return data_set


def getQuality():

    global selected_quality_option
    selected_q_option = selected_quality_option.get()
    selected_f_option = selected_fontsize_option.get()
    titles_font, legend_font = fontsize_options[selected_f_option]
    width, height = quality_options[selected_q_option]
    return width, height, titles_font, legend_font


# fool proofing

def show_error_message(message):
    messagebox.showerror("Error", message)


def try_file_path(file_path):
    if os.path.exists(file_path):
        return True
    show_error_message("File dose not exist")
    return False


def try_sheet(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            return True
        else:
            show_error_message("Sheet dose not exist")
            return False
    except FileNotFoundError:
        return False


def validate_entries(entry, name):
    text1 = entry.get().strip()
    if not text1:
        show_error_message(f"{name} must be filled")
        return False
    else:
        return True


def plotcommands(data_set, plot_type):
    if plot_type == "Bar":
        figure = barplot(data_set)
    if plot_type == "Box":
        figure = boxplot(data_set)
    if plot_type == "Strip":
        figure = stripplot(data_set)
    if plot_type == "Pval":
        data_set.p_value = tukey_kermer(data_set)
        if data_set.p_value is None:
            message = "P-values not available. Please generate them first."
            show_error_message(message)
            return
        printmatrix(data_set.p_value, data_set)
        return
    exportfigure(
        figure, f"{data_set.export_folder} {data_set.title}_{plot_type} Plot.png")
    return


# Button commands


def handle_button(plot_type):
    global current_year
    if (current_year > licenseing()+4):
        message = "License has expired contact creator"
        show_error_message(message)
    else:
        if validate_entries(file_path_entry, "File path") \
                and validate_entries(sheet_name_entry, "Sheet name")\
                and validate_entries(x_axis_entry, "X Axis Lable ")\
                and validate_entries(y_axis_entry, "Y Axis Lable ")\
                and validate_entries(title_entry, "Title ")\
                and validate_entries(export_folder_entry, "Export folder"):
            if try_file_path(file_path_entry.get()) and try_sheet(file_path_entry.get(), sheet_name_entry.get()):
                data_set = DataSetConstructior(file_path_entry.get(),
                                               sheet_name_entry.get(),
                                               x_axis_entry.get(),
                                               y_axis_entry.get(),
                                               title_entry.get(),
                                               export_folder_entry.get())
            plotcommands(data_set, plot_type)

    return


# Make GUI
# Tkinter window setup
root = tk.Tk()
root.title("Graph Maker")
root.geometry("1260x330")
style = ttk.Style()
style.theme_use("xpnative")
style.configure("TButton", font=("Arial", 14))
style.configure("TLabel", font=("Arial", 14))
style.configure("TMenubutton", font=("Arial", 14))
selected_quality_option = tk.StringVar()
selected_quality_option.set("800 X 600")
quality_options = {
    "800 X 600": (800, 600),
    "960 X 720": (960, 720),
    "1200 X 900": (1200, 900),
    "1440 X 1080": (1440, 1080),
    "1600 X 1200": (1600, 1200)
}

selected_fontsize_option = tk.StringVar()
selected_fontsize_option.set("14 X 12")
fontsize_options = {
    "14 X 12": (14, 12),
    "16 X 14": (16, 14),
    "21 X 18": (21, 18),
    "25 X 21": (25, 21),
    "28 X 24": (28, 24),
    "32 X 28": (32, 28),
    "36 X 32": (36, 32),
    "40 X 36": (40, 36)
}

# Make Widgets

# lables
Spacer = ttk.Label(
    root, text="  ")
Mainlabel = ttk.Label(root, text="Graph Maker", font=("Arial", 18, "bold"))
file_path_label = ttk.Label(
    root, text="Enter the location of you Exel file")
sheet_name_label = ttk.Label(
    root, text="Enter the name of the sheet")
x_axis_label = ttk.Label(
    root, text="Enter your X Axis Lable")
y_axis_label = ttk.Label(
    root, text="Enter your Y Axis Lable")
title_label = ttk.Label(
    root, text="Enter the title of your graph")
export_folder_label = ttk.Label(
    root, text="Enter the location to export")
quality_option_menu__label = ttk.Label(
    root, text="Quality")
fontsize_option_menu__label = ttk.Label(
    root, text="Font Size")
# Entries
file_path_entry = ttk.Entry(root, font=("Arial", 14), width=50)
sheet_name_entry = ttk.Entry(root, font=("Arial", 14), width=50)
x_axis_entry = ttk.Entry(root, font=("Arial", 14), width=50)
y_axis_entry = ttk.Entry(root, font=("Arial", 14), width=50)
title_entry = ttk.Entry(root, font=("Arial", 14), width=50)
export_folder_entry = ttk.Entry(root, font=("Arial", 14), width=50)
quality_option_menu = ttk.OptionMenu(
    root, selected_quality_option, selected_quality_option.get(), *quality_options.keys())
fontsize_option_menu = ttk.OptionMenu(
    root, selected_fontsize_option, selected_fontsize_option.get(), *fontsize_options.keys())
# Buttons
bar_plot_button = ttk.Button(
    root, text="Bar Graph", command=lambda: handle_button("Bar"))
box_plot_button = ttk.Button(
    root, text="Box Plot", command=lambda: handle_button("Box"))
strip_plot_button = ttk.Button(
    root, text="Strip Plot", command=lambda: handle_button("Strip"))
p_value_button = ttk.Button(
    root, text="Export p-Value table", command=lambda:  handle_button("Pval"))


# Place Widgets
Spacer.grid(row=0, column=0)
counterCol = 1
counterRow = 0
Mainlabel.grid(row=counterRow, column=counterCol+1, columnspan=4)
counterCol += 2
counterRow += 1
file_path_label.grid(row=counterRow, column=counterCol)
file_path_entry.grid(row=counterRow+1, column=counterCol)
counterCol += 2
sheet_name_label.grid(row=counterRow, column=counterCol)
sheet_name_entry.grid(row=counterRow+1, column=counterCol)
counterCol -= 2
counterRow += 2
x_axis_label.grid(row=counterRow, column=counterCol)
x_axis_entry.grid(row=counterRow+1, column=counterCol)
counterCol += 2
y_axis_label.grid(row=counterRow, column=counterCol)
y_axis_entry.grid(row=counterRow+1, column=counterCol)
counterCol -= 2
counterRow += 2
title_label.grid(row=counterRow, column=counterCol)
title_entry.grid(row=counterRow+1, column=counterCol)
counterCol += 2
export_folder_label.grid(row=counterRow, column=counterCol)
export_folder_entry.grid(row=counterRow+1, column=counterCol)
counterCol -= 1
counterRow += 2
quality_option_menu__label.grid(row=counterRow, column=counterCol)
quality_option_menu.grid(row=counterRow+1, column=counterCol)
counterRow += 2
fontsize_option_menu__label.grid(row=counterRow, column=counterCol)
fontsize_option_menu.grid(row=counterRow+1, column=counterCol)
counterCol -= 1
bar_plot_button.grid(row=counterRow, column=counterCol)
counterCol += 2
box_plot_button.grid(row=counterRow, column=counterCol)
counterCol -= 2
counterRow += 1
strip_plot_button.grid(row=counterRow, column=counterCol)
counterCol += 2
p_value_button.grid(row=counterRow, column=counterCol)

# Run system
root.mainloop()
